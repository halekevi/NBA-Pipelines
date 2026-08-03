#!/usr/bin/env python3
"""
Backtest PrizePicks line movement across scheduled refresh runs.

Uses preserved combined_slate_tickets_*.bak_<timestamp>.xlsx files under outputs/
(each refresh with -NoOverwrite backs up the prior slate). Compares lines at
9 AM / 11 AM / 1 PM windows (legacy 10:30 AM hour=10 included when present).

Joins graded_props JSON for the slate date to score which run had the most
favorable line for each direction (OVER: lower line, UNDER: higher line).

Usage:
  py -3 scripts/backtest_pp_line_moves.py
  py -3 scripts/backtest_pp_line_moves.py --from 2026-05-07 --standard-only
  py -3 scripts/backtest_pp_line_moves.py --csv data/reports/pp_line_move_backtest.csv
"""

from __future__ import annotations

import argparse
import hashlib
import json
import pickle
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

_REPO = Path(__file__).resolve().parent.parent
_SCRIPTS = _REPO / "scripts"
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from analyze_graded_history import (  # noqa: E402
    _GRADED_DIR,
    _norm_dir,
    _norm_pick,
    _norm_sport,
    _parse_hit,
)

_BAK_RE = re.compile(
    r"combined_slate_tickets_(?P<slate>\d{4}-\d{2}-\d{2})\.bak_(?P<stamp>\d{8}_\d{6})\.(?P<ext>xlsx|json)$",
    re.I,
)
_CACHE_DIR = _REPO / "data" / "cache" / "pp_line_move_slates"

# Ordered windows from first fetch → midday refreshes.
# 5AM/8AM = early board (Daily 5AM / Daily 8AM bak stamps); 11AM historical only.
RUN_ORDER = ("5AM", "8AM", "9AM", "1030AM", "11AM", "1PM")


def _norm(raw: Any) -> str:
    return " ".join(str(raw or "").strip().lower().split())


def _to_float(raw: Any) -> float | None:
    try:
        if raw is None or str(raw).strip() == "":
            return None
        return float(raw)
    except (TypeError, ValueError):
        return None


def _hour_to_run(hour: int) -> str | None:
    # Bak stamp is usually when the prior slate was copied at refresh start.
    if hour in (4, 5, 6, 7):
        return "5AM"
    if hour == 8:
        return "8AM"
    if hour == 9:
        return "9AM"
    if hour == 10:
        return "1030AM"
    if hour == 11:
        return "11AM"
    if hour in (12, 13, 14):
        return "1PM"
    return None


@dataclass(frozen=True)
class LegKey:
    sport: str
    player: str
    prop: str
    direction: str
    pick_type: str
    game_date: str


def _cache_path_for(source: Path) -> Path:
    st = source.stat()
    digest = hashlib.sha1(f"{source.resolve()}|{st.st_mtime_ns}|{st.st_size}".encode()).hexdigest()[:16]
    return _CACHE_DIR / f"{digest}.pkl"


def _legkey_tuple(key: LegKey) -> tuple[str, ...]:
    return (key.sport, key.player, key.prop, key.direction, key.pick_type, key.game_date)


def _legkey_from_tuple(t: tuple[str, ...]) -> LegKey:
    return LegKey(t[0], t[1], t[2], t[3], t[4], t[5] if len(t) > 5 else "")


def _load_cached_slate(source: Path) -> dict[LegKey, dict[str, Any]] | None:
    cp = _cache_path_for(source)
    if not cp.is_file():
        return None
    try:
        blob = pickle.loads(cp.read_bytes())
        if blob.get("source_mtime_ns") != source.stat().st_mtime_ns:
            return None
        return {_legkey_from_tuple(k): v for k, v in blob["rows"].items()}
    except Exception:
        return None


def _save_cached_slate(source: Path, rows: dict[LegKey, dict[str, Any]]) -> None:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cp = _cache_path_for(source)
    blob = {
        "source_mtime_ns": source.stat().st_mtime_ns,
        "rows": {_legkey_tuple(k): v for k, v in rows.items()},
    }
    cp.write_bytes(pickle.dumps(blob, protocol=pickle.HIGHEST_PROTOCOL))


def _rows_from_xlsx(path: Path) -> dict[LegKey, dict[str, Any]]:
    base_cols = ["Sport", "Player", "Prop", "Dir", "Pick Type", "Line", "Game Date"]
    optional_cols = ["Standard Line"]
    # Some older slates omit Standard Line; read available columns only.
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Full Slate" not in wb.sheetnames:
            return {}
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(wb["Full Slate"].iter_rows(min_row=1, max_row=1))
        ]
    finally:
        wb.close()
    usecols = [c for c in base_cols + optional_cols if c in headers]
    df = pd.read_excel(path, sheet_name="Full Slate", usecols=usecols, engine="openpyxl")
    out: dict[LegKey, dict[str, Any]] = {}
    for r in df.itertuples(index=False):
        sport = _norm_sport(getattr(r, "Sport", ""))
        player = _norm(getattr(r, "Player", ""))
        prop = _norm(getattr(r, "Prop", ""))
        direction = _norm_dir(getattr(r, "Dir", ""))
        pick_type = _norm_pick(getattr(r, "Pick_Type", getattr(r, "Pick Type", ""))).title()
        if not sport or not player or not prop or direction not in ("OVER", "UNDER"):
            continue
        gd = getattr(r, "Game_Date", getattr(r, "Game Date", ""))
        game_date = str(gd or "").strip()[:10]
        key = LegKey(sport, player, prop, direction, pick_type, game_date)
        out[key] = {
            "line": _to_float(getattr(r, "Line", None)),
            "standard_line": _to_float(
                getattr(r, "Standard_Line", getattr(r, "Standard Line", None))
            ),
        }
    return out


def _rows_from_json(path: Path) -> dict[LegKey, dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    out: dict[LegKey, dict[str, Any]] = {}
    for group in payload.get("groups") or []:
        for ticket in group.get("tickets") or []:
            for leg in ticket.get("legs") or []:
                sport = _norm_sport(leg.get("sport"))
                player = _norm(leg.get("player"))
                prop = _norm(leg.get("prop_type") or leg.get("prop"))
                direction = _norm_dir(leg.get("direction") or leg.get("dir"))
                pick_type = _norm_pick(leg.get("pick_type")).title()
                if not sport or not player or not prop or direction not in ("OVER", "UNDER"):
                    continue
                game_date = str(leg.get("game_date") or "").strip()[:10]
                key = LegKey(sport, player, prop, direction, pick_type, game_date)
                out[key] = {
                    "line": _to_float(leg.get("line")),
                    "standard_line": _to_float(leg.get("standard_line")),
                }
    return out


def _read_full_slate(path: Path) -> dict[LegKey, dict[str, Any]]:
    cached = _load_cached_slate(path)
    if cached is not None:
        return cached
    try:
        if path.suffix.lower() == ".json":
            rows = _rows_from_json(path)
        else:
            rows = _rows_from_xlsx(path)
    except Exception as exc:
        print(f"  [warn] skip unreadable slate: {path.name} ({exc})", file=sys.stderr)
        return {}
    if rows:
        _save_cached_slate(path, rows)
    return rows


def _discover_backups(
    outputs_dir: Path,
    json_backups_dir: Path,
    min_date: str,
    extra_outputs: list[Path] | None = None,
) -> list[dict[str, Any]]:
    """Prefer xlsx (full slate); fall back to json ticket legs for same stamp."""
    by_stamp: dict[tuple[str, str], dict[str, Any]] = {}
    search_roots: list[tuple[Path, str]] = [
        (outputs_dir, ".xlsx"),
        (json_backups_dir, ".json"),
    ]
    for extra in extra_outputs or []:
        if extra.is_dir():
            search_roots.append((extra, ".xlsx"))
    for root, ext in search_roots:
        if not root.is_dir():
            continue
        pattern = f"**/combined_slate_tickets_*.bak_*{ext}"
        for path in sorted(root.glob(pattern)):
            m = _BAK_RE.search(path.name)
            if not m:
                continue
            slate_date = m.group("slate")
            if slate_date < min_date:
                continue
            stamp = m.group("stamp")
            dt = datetime.strptime(stamp, "%Y%m%d_%H%M%S")
            run = _hour_to_run(dt.hour)
            if run is None:
                continue
            key = (slate_date, stamp)
            entry = {
                "path": path,
                "slate_date": slate_date,
                "run": run,
                "captured_at": dt.isoformat(sep=" "),
                "hour": dt.hour,
                "source": "xlsx" if ext == ".xlsx" else "json",
            }
            existing = by_stamp.get(key)
            if existing is None or (existing["source"] == "json" and entry["source"] == "xlsx"):
                by_stamp[key] = entry
    return list(by_stamp.values())


def _pick_one_per_run(backups: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Earliest backup in each run window per slate date."""
    by_run: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for b in backups:
        by_run[b["run"]].append(b)
    out: dict[str, dict[str, Any]] = {}
    for run, items in by_run.items():
        out[run] = min(items, key=lambda x: x["path"].name)
    return out


def _load_graded_for_date(slate_date: str) -> dict[tuple, dict[str, Any]]:
    path = _GRADED_DIR / f"graded_props_{slate_date}.json"
    if not path.is_file():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    rows = payload.get("props") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        return {}

    out: dict[tuple, dict[str, Any]] = {}
    for r in rows:
        sport = _norm_sport(r.get("sport"))
        player = _norm(r.get("player"))
        prop = _norm(r.get("prop") or r.get("prop_type") or r.get("category"))
        direction = _norm_dir(r.get("direction"))
        pick_type = _norm_pick(r.get("pick_type")).title()
        if not sport or not player or not prop:
            continue
        hit = _parse_hit(r.get("result"))
        actual = _to_float(r.get("actual_value"))
        key = (sport, player, prop, direction, pick_type)
        out[key] = {
            "hit": hit,
            "actual": actual,
            "line": _to_float(r.get("line")),
            "result": str(r.get("result") or "").strip().upper(),
        }
    return out


def _would_hit(actual: float, line: float, direction: str) -> int | None:
    if direction == "OVER":
        if actual > line:
            return 1
        if actual < line:
            return 0
        return None  # push
    if direction == "UNDER":
        if actual < line:
            return 1
        if actual > line:
            return 0
        return None
    return None


def _favorability(line: float, direction: str) -> float:
    """Higher = more favorable to the bettor."""
    return -line if direction == "OVER" else line


def _compare_pair(
    old_run: str,
    new_run: str,
    old_rows: dict[LegKey, dict[str, Any]],
    new_rows: dict[LegKey, dict[str, Any]],
    *,
    standard_only: bool,
) -> list[dict[str, Any]]:
    moves: list[dict[str, Any]] = []
    shared = set(old_rows.keys()) & set(new_rows.keys())
    for key in shared:
        if standard_only and key.pick_type.lower() != "standard":
            continue
        a = old_rows[key]
        b = new_rows[key]
        old_line = a.get("line")
        new_line = b.get("line")
        if old_line is None or new_line is None or old_line == new_line:
            continue
        delta = new_line - old_line
        moves.append(
            {
                "sport": key.sport,
                "player": key.player,
                "prop": key.prop,
                "direction": key.direction,
                "pick_type": key.pick_type,
                "game_date": key.game_date,
                "old_run": old_run,
                "new_run": new_run,
                "old_line": old_line,
                "new_line": new_line,
                "line_delta": delta,
                "abs_delta": abs(delta),
                "moved_favorable": (
                    (key.direction == "OVER" and delta < 0)
                    or (key.direction == "UNDER" and delta > 0)
                ),
            }
        )
    return moves


def _score_runs_for_leg(
    timeline: dict[str, float],
    direction: str,
    graded: dict[str, Any] | None,
) -> dict[str, Any]:
    actual = graded.get("actual") if graded else None
    official_hit = graded.get("hit") if graded else None
    run_hits: dict[str, int | None] = {}
    run_fav: dict[str, float] = {}
    for run, line in timeline.items():
        run_fav[run] = _favorability(line, direction)
        if actual is not None:
            run_hits[run] = _would_hit(actual, line, direction)

    best_fav_run = max(timeline.keys(), key=lambda r: run_fav[r]) if timeline else None
    best_hit_run = None
    if actual is not None:
        hitting = [r for r, h in run_hits.items() if h == 1]
        if hitting:
            best_hit_run = max(hitting, key=lambda r: run_fav[r])
        else:
            def miss_margin(r: str) -> float:
                line = timeline[r]
                if direction == "OVER":
                    return line - actual
                return actual - line

            best_hit_run = min(timeline.keys(), key=miss_margin)

    first_run = next((r for r in RUN_ORDER if r in timeline), None)
    last_run = next((r for r in reversed(RUN_ORDER) if r in timeline), None)

    return {
        "best_favorable_run": best_fav_run,
        "best_hit_run": best_hit_run,
        "first_run": first_run,
        "last_run": last_run,
        "run_hits": run_hits,
        "run_lines": timeline,
        "official_hit": official_hit,
        "actual": actual,
        "line_moved": len({v for v in timeline.values()}) > 1,
        "first_vs_last_delta": (
            timeline[last_run] - timeline[first_run]
            if first_run and last_run and first_run != last_run
            else 0.0
        ),
    }


def run_backtest(
    *,
    min_date: str,
    standard_only: bool,
    outputs_dir: Path,
    json_backups_dir: Path,
    extra_outputs: list[Path] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    backups = _discover_backups(
        outputs_dir, json_backups_dir, min_date, extra_outputs=extra_outputs
    )
    by_date: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for b in backups:
        by_date[b["slate_date"]].append(b)

    transition_rows: list[dict[str, Any]] = []
    leg_rows: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []

    for slate_date in sorted(by_date.keys()):
        per_run = _pick_one_per_run(by_date[slate_date])
        runs_present = [r for r in RUN_ORDER if r in per_run]
        if len(runs_present) < 2:
            continue

        slate_by_run: dict[str, dict[LegKey, dict[str, Any]]] = {}
        for run in runs_present:
            slate_by_run[run] = _read_full_slate(per_run[run]["path"])
        runs_present = [r for r in runs_present if slate_by_run.get(r)]
        if len(runs_present) < 2:
            continue

        day_moves = 0
        day_abs = 0.0
        day_fav = 0
        for i in range(len(runs_present) - 1):
            old_run = runs_present[i]
            new_run = runs_present[i + 1]
            pair_moves = _compare_pair(
                old_run,
                new_run,
                slate_by_run[old_run],
                slate_by_run[new_run],
                standard_only=standard_only,
            )
            for m in pair_moves:
                m["slate_date"] = slate_date
                transition_rows.append(m)
            day_moves += len(pair_moves)
            day_abs += sum(m["abs_delta"] for m in pair_moves)
            day_fav += sum(1 for m in pair_moves if m["moved_favorable"])

        daily_rows.append(
            {
                "slate_date": slate_date,
                "runs": ",".join(runs_present),
                "line_moves": day_moves,
                "abs_delta_sum": round(day_abs, 3),
                "favorable_moves": day_fav,
                "favorable_move_pct": round(100 * day_fav / day_moves, 1) if day_moves else None,
            }
        )

        graded = _load_graded_for_date(slate_date)
        # Build per-leg timelines across runs
        all_keys: set[LegKey] = set()
        for run in runs_present:
            all_keys.update(slate_by_run[run].keys())

        for key in all_keys:
            if standard_only and key.pick_type.lower() != "standard":
                continue
            timeline: dict[str, float] = {}
            for run in runs_present:
                row = slate_by_run[run].get(key)
                if row and row.get("line") is not None:
                    timeline[run] = float(row["line"])
            if len(timeline) < 2:
                continue

            gkey = (key.sport, key.player, key.prop, key.direction, key.pick_type)
            scored = _score_runs_for_leg(timeline, key.direction, graded.get(gkey))
            if scored["actual"] is None and scored["official_hit"] is None:
                continue

            leg_rows.append(
                {
                    "slate_date": slate_date,
                    "sport": key.sport,
                    "player": key.player,
                    "prop": key.prop,
                    "direction": key.direction,
                    "pick_type": key.pick_type,
                    "runs_present": ",".join(timeline.keys()),
                    "official_hit": scored["official_hit"],
                    "actual": scored["actual"],
                    "best_favorable_run": scored["best_favorable_run"],
                    "best_hit_run": scored["best_hit_run"],
                    "first_run": scored["first_run"],
                    "last_run": scored["last_run"],
                    "first_vs_last_delta": scored["first_vs_last_delta"],
                    **{f"line_{r}": timeline.get(r) for r in RUN_ORDER},
                    **{f"hit_{r}": scored["run_hits"].get(r) for r in RUN_ORDER},
                }
            )

    transitions = pd.DataFrame(transition_rows)
    legs = pd.DataFrame(leg_rows)
    daily = pd.DataFrame(daily_rows)

    summary_rows: list[dict[str, Any]] = []
    if not transitions.empty:
        for (old_run, new_run), g in transitions.groupby(["old_run", "new_run"], sort=False):
            summary_rows.append(
                {
                    "transition": f"{old_run} → {new_run}",
                    "old_run": old_run,
                    "new_run": new_run,
                    "move_count": len(g),
                    "avg_abs_delta": round(g["abs_delta"].mean(), 3),
                    "median_abs_delta": round(g["abs_delta"].median(), 3),
                    "pct_favorable": round(100 * g["moved_favorable"].mean(), 1),
                    "pct_unfavorable": round(100 * (1.0 - g["moved_favorable"].mean()), 1),
                    "favorable_n": int(g["moved_favorable"].sum()),
                    "unfavorable_n": int((~g["moved_favorable"]).sum()),
                    "fav_abs_sum": round(
                        g.loc[g["moved_favorable"], "abs_delta"].sum(), 2
                    ),
                    "unfav_abs_sum": round(
                        g.loc[~g["moved_favorable"], "abs_delta"].sum(), 2
                    ),
                    "slate_days": g["slate_date"].nunique(),
                }
            )
        by_sport = (
            transitions.groupby(["old_run", "new_run", "sport"], dropna=False)
            .agg(move_count=("abs_delta", "size"), avg_abs_delta=("abs_delta", "mean"))
            .reset_index()
        )
    else:
        by_sport = pd.DataFrame()

    summary = pd.DataFrame(summary_rows)
    if not summary.empty:
        summary = summary.sort_values("move_count", ascending=False)

    return summary, transitions, legs, daily, by_sport


def _print_hit_timing(legs: pd.DataFrame) -> None:
    if legs.empty:
        print("\nNo graded legs with multi-run lines found.")
        return

    print("\n=== Hit rate by run window (legs with actuals) ===")
    for run in RUN_ORDER:
        col = f"hit_{run}"
        if col not in legs.columns:
            continue
        sub = legs[legs[col].notna()]
        if sub.empty:
            continue
        hr = 100 * sub[col].mean()
        print(f"  {run:6s}  {hr:5.1f}% hit  (n={len(sub):,})")

    print("\n=== Best time to pick (most favorable line that still hits) ===")
    vc = legs["best_hit_run"].value_counts(dropna=True)
    total = vc.sum()
    for run, n in vc.items():
        print(f"  {run:6s}  {100 * n / total:5.1f}%  (n={int(n):,})")

    print("\n=== Best favorable line (regardless of hit) ===")
    vc2 = legs["best_favorable_run"].value_counts(dropna=True)
    total2 = vc2.sum()
    for run, n in vc2.items():
        print(f"  {run:6s}  {100 * n / total2:5.1f}%  (n={int(n):,})")

    improved = legs[
        legs["best_hit_run"].notna()
        & legs["hit_9AM"].notna()
        & (legs["best_hit_run"] != legs["first_run"])
    ]
    if not improved.empty and "hit_9AM" in legs.columns:
        saved = 0
        for _, r in improved.iterrows():
            first = r["first_run"]
            best = r["best_hit_run"]
            if pd.notna(r.get(f"hit_{first}")) and pd.notna(r.get(f"hit_{best}")):
                if r[f"hit_{first}"] == 0 and r[f"hit_{best}"] == 1:
                    saved += 1
        print(f"\n  Legs where waiting beat first snapshot: {saved:,} / {len(improved):,}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Backtest PP line moves across refresh runs.")
    ap.add_argument("--from", dest="min_date", default="2026-05-01", help="Min slate date YYYY-MM-DD")
    ap.add_argument("--outputs", default=str(_REPO / "outputs"), help="outputs/ root")
    ap.add_argument(
        "--extra-outputs",
        nargs="*",
        default=[],
        help="Additional outputs/ roots to scan for bak xlsx",
    )
    ap.add_argument(
        "--json-backups",
        default=str(_REPO / "ui_runner" / "data" / "backups"),
        help="ui_runner JSON slate backups",
    )
    ap.add_argument("--standard-only", action="store_true", help="Standard pick_type only")
    ap.add_argument("--csv", default="", help="Write leg-level CSV to this path")
    args = ap.parse_args()

    extra = [Path(p).expanduser().resolve() for p in (args.extra_outputs or [])]
    summary, transitions, legs, daily, by_sport = run_backtest(
        min_date=str(args.min_date).strip()[:10],
        standard_only=bool(args.standard_only),
        outputs_dir=Path(args.outputs).expanduser().resolve(),
        json_backups_dir=Path(args.json_backups).expanduser().resolve(),
        extra_outputs=extra,
    )

    n_days = daily["slate_date"].nunique() if not daily.empty else 0
    print(f"PP line-move backtest  (slate days={n_days}, from={args.min_date})")
    if args.standard_only:
        print("  filter: Standard pick_type only")

    if summary.empty:
        print("\nNo line moves found in backup slates. Ensure outputs/*/combined_slate_tickets_*.bak_*.xlsx exist.")
        return 1

    print("\n=== Movement volume by transition ===")
    print(summary.to_string(index=False))

    if not transitions.empty:
        print("\n=== Favorable vs unfavorable (Standard line vs pick direction) ===")
        print("  Favorable: OVER line down / UNDER line up. Unfavorable: opposite.")
        rows = []
        for new_run, g in transitions.groupby("new_run", sort=False):
            fav = g[g["moved_favorable"]]
            unfav = g[~g["moved_favorable"]]
            rows.append(
                {
                    "into_window": new_run,
                    "moves": len(g),
                    "fav_n": len(fav),
                    "unfav_n": len(unfav),
                    "fav_pct": round(100 * len(fav) / len(g), 1),
                    "unfav_pct": round(100 * len(unfav) / len(g), 1),
                    "avg_abs": round(g["abs_delta"].mean(), 3),
                    "fav_abs_sum": round(fav["abs_delta"].sum(), 1),
                    "unfav_abs_sum": round(unfav["abs_delta"].sum(), 1),
                    "net_fav_abs": round(fav["abs_delta"].sum() - unfav["abs_delta"].sum(), 1),
                }
            )
        dest_df = pd.DataFrame(rows)
        order = {r: i for i, r in enumerate(RUN_ORDER)}
        dest_df["_o"] = dest_df["into_window"].map(lambda x: order.get(x, 99))
        dest_df = dest_df.sort_values("_o").drop(columns="_o")
        print(dest_df.to_string(index=False))

        # Cumulative first→later from adjacent chain isn't enough; recompute from legs
        # when present, else skip (full-slate first→later added in run_backtest reports).
        print("\n=== First snapshot → later windows (graded legs with multi-run lines) ===")
        first_later_rows = []
        if not legs.empty:
            for later in RUN_ORDER[1:]:
                col = f"line_{later}"
                first_cols = [f"line_{r}" for r in RUN_ORDER]
                sub = legs[legs[col].notna()].copy()
                if sub.empty:
                    continue
                deltas = []
                favs = []
                for _, r in sub.iterrows():
                    first_line = None
                    for fc in first_cols:
                        if fc == col:
                            break
                        v = r.get(fc)
                        if pd.notna(v):
                            first_line = float(v)
                            break
                    if first_line is None:
                        continue
                    cur = float(r[col])
                    if abs(cur - first_line) < 1e-9:
                        continue
                    d = cur - first_line
                    direction = str(r["direction"])
                    fav = (direction == "OVER" and d < 0) or (direction == "UNDER" and d > 0)
                    deltas.append(abs(d))
                    favs.append(fav)
                if not deltas:
                    continue
                fav_n = sum(1 for x in favs if x)
                unfav_n = len(favs) - fav_n
                first_later_rows.append(
                    {
                        "vs_first→": later,
                        "moved_legs": len(deltas),
                        "fav_n": fav_n,
                        "unfav_n": unfav_n,
                        "fav_pct": round(100 * fav_n / len(deltas), 1),
                        "unfav_pct": round(100 * unfav_n / len(deltas), 1),
                        "avg_abs": round(sum(deltas) / len(deltas), 3),
                    }
                )
            if first_later_rows:
                print(pd.DataFrame(first_later_rows).to_string(index=False))
            else:
                print("  (no graded multi-run moved legs; use adjacent transitions above)")
        else:
            print("  (no graded legs; adjacent transitions above are the main signal)")

    if not by_sport.empty:
        top = by_sport.sort_values("move_count", ascending=False).head(12)
        print("\n=== Top sport × transition (by move count) ===")
        print(top.to_string(index=False))

    if not daily.empty:
        print("\n=== Busiest slate days (total line moves) ===")
        busy = daily.sort_values("line_moves", ascending=False).head(8)
        print(busy.to_string(index=False))

    _print_hit_timing(legs)

    if args.csv:
        out = Path(args.csv).expanduser().resolve()
        out.parent.mkdir(parents=True, exist_ok=True)
        legs.to_csv(out, index=False)
        print(f"\nWrote leg detail: {out}")

    report_dir = _REPO / "data" / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    summary.to_csv(report_dir / "pp_line_move_summary.csv", index=False)
    transitions.to_csv(report_dir / "pp_line_move_transitions.csv", index=False)
    legs.to_csv(report_dir / "pp_line_move_legs.csv", index=False)
    daily.to_csv(report_dir / "pp_line_move_daily.csv", index=False)
    print(f"\nReports: {report_dir}/pp_line_move_*.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
